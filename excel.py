import os

from openpyxl import Workbook, load_workbook
from openpyxl import load_workbook

if not os.path.isdir('input'):
    os.mkdir('input')
if not os.path.isdir('output'):
    os.mkdir('output')

isInitialized = False


# 시작 시 반드시 init()을 호출해야 함
def init(_filename: str):
    global wb_input, wb_output, isInitialized, filename
    wb_input = load_workbook(f'input/{_filename}')
    wb_output = Workbook()
    filename = _filename
    isInitialized = True

# 종료 전 반드시 save()를 호출해야 함
def save():
    global wb_input, wb_output, isInitialized
    if not isInitialized:
        print('Error: excel.init()을 먼저 호출해주세요.')
        exit()
    wb_output.save(f'output/{filename}')

# input 엑셀 파일의 셀에서 값을 읽어서 반환
def read(sheetname: str, row: int, column: str):
    global wb_input, wb_output, isInitialized
    if not isInitialized:
        print('Error: excel.init()을 먼저 호출해주세요.')
        exit()
    ws = wb_input[sheetname]
    write(sheetname, row, column, ws[column + str(row)].value)
    return ws[column + str(row)].value

# output 엑셀 파일의 셀에 값을 씀
def write(sheetname: str, row: int, column: str, value: str):
    global wb_input, wb_output, isInitialized
    if not isInitialized:
        print('Error: excel.init()을 먼저 호출해주세요.')
        exit()
    if sheetname in wb_output.sheetnames:
        ws = wb_output[sheetname]
    else:
        ws = wb_output.create_sheet(sheetname)
    ws[column + str(row)] = value
